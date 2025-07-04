import pandas as pd
import streamlit as st
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

st.set_page_config(page_title="Calculadora de Pedido SUOLMEX", layout="wide")

# ESTILO
st.markdown("""
    <style>
        body {
            background-color: #f5f7fa;
        }
        .stApp { font-family: 'Segoe UI', sans-serif; }
        h1, h2, h3 { color: #00264d; }
        .stButton>button {
            background-color: #00264d;
            color: white;
            font-weight: bold;
            padding: 8px 20px;
            border-radius: 6px;
            border: none;
        }
        .stButton>button:hover { background-color: #003366; }
        .stDataFrame thead tr th {
            background-color: #e6e9ef;
            color: #00264d;
        }
    </style>
""", unsafe_allow_html=True)

# SIDEBAR
with st.sidebar:
    from pathlib import Path
    logo_path = Path(__file__).parent / "logo_suolmex.jpg"
    st.image(logo_path, width=200)
    st.markdown("### Instrucciones")
    st.markdown("""
    1. Sube un archivo Excel con el pedido o ingrésalo manualmente.
    2. Verifica el resumen del pedido.
    3. Exporta a Excel al finalizar.
    4. Puedes reiniciar para comenzar de nuevo.
    """)

@st.cache_data
def cargar_fichas():
    excel_file = pd.ExcelFile("FICHAS2.xlsx")
    hojas_deseadas = ['6001', '2066', '2060', '4098', 'PLANTILLAS']
    dataframes = []
    for hoja in hojas_deseadas:
        if hoja in excel_file.sheet_names:
            df_temp = excel_file.parse(hoja)
            df_temp['Hoja'] = hoja
            dataframes.append(df_temp)
    df = pd.concat(dataframes, ignore_index=True)
    df["Codigo del Producto"] = df["Codigo del Producto"].astype(str).str.strip()
    df["Linea"] = df["Linea"].astype(str).str.strip().str.upper()
    df["Corrida"] = df["Corrida"].astype(str).str.strip()
    df["Peso/Pie"] = pd.to_numeric(df["Peso/Pie"], errors="coerce")
    return df.dropna(subset=["Peso/Pie", "Relacion Poliol:ISO"])

fichas = cargar_fichas()

if "pedido_total" not in st.session_state:
    st.session_state["pedido_total"] = []
if "modelo_actual" not in st.session_state:
    st.session_state["modelo_actual"] = ""
if "corrida_seleccionada" not in st.session_state:
    st.session_state["corrida_seleccionada"] = None

st.subheader("Subir archivo de pedido (opcional)")
archivo_pedido = st.file_uploader("Selecciona un archivo Excel con el pedido", type=["xlsx"])
if archivo_pedido:
    pedido_df = pd.read_excel(archivo_pedido)
    st.success("Pedido importado correctamente.")
    for _, row in pedido_df.iterrows():
        codigo = str(row["Codigo del Producto"]).strip()
        modelo = str(row["Modelo"]).strip().upper()
        corrida = str(row["Talla"]).strip()
        cantidad = int(row["Cantidad pares"])
        ficha = fichas[(fichas["Codigo del Producto"] == codigo) &
                       (fichas["Linea"] == modelo) &
                       (fichas["Corrida"] == corrida)]
        if not ficha.empty:
            ficha = ficha.iloc[0]
            peso_total = ficha['Peso/Pie'] * cantidad * 2
            poliol, iso = map(float, ficha['Relacion Poliol:ISO'].split(":"))
            total_partes = poliol + iso
            cantidad_poliol = peso_total * (poliol / total_partes)
            cantidad_iso = peso_total * (iso / total_partes)
            st.session_state["pedido_total"].append({
                "Código": codigo,
                "Modelo": modelo,
                "Talla": corrida,
                "Cantidad pares": cantidad,
                "Peso Total (g)": peso_total,
                "Poliol (g)": cantidad_poliol,
                "ISO (g)": cantidad_iso,
                "Hoja": ficha['Hoja']
            })

st.markdown("---")
st.subheader("Ingreso manual")
codigo = st.selectbox("Código del Producto:", sorted(fichas["Codigo del Producto"].unique()))
modelos = fichas[fichas["Codigo del Producto"] == codigo]["Linea"].unique()
modelo = st.selectbox("Modelo:", sorted(modelos))
cantidad = st.number_input("Cantidad de pares:", min_value=1, step=1)

corridas = fichas[(fichas["Codigo del Producto"] == codigo) & (fichas["Linea"] == modelo)]["Corrida"].unique()
st.markdown("#### Selecciona una talla:")
cols = st.columns(min(5, len(corridas)))
for i, talla in enumerate(sorted(corridas)):
    if cols[i % 5].button(talla):
        st.session_state["corrida_seleccionada"] = talla

if st.session_state["corrida_seleccionada"]:
    corrida = st.session_state["corrida_seleccionada"]
    ficha = fichas[(fichas["Codigo del Producto"] == codigo) &
                   (fichas["Linea"] == modelo) &
                   (fichas["Corrida"] == corrida)]
    if not ficha.empty:
        ficha = ficha.iloc[0]
        peso_total = ficha['Peso/Pie'] * cantidad * 2
        poliol, iso = map(float, ficha['Relacion Poliol:ISO'].split(":"))
        total_partes = poliol + iso
        cantidad_poliol = peso_total * (poliol / total_partes)
        cantidad_iso = peso_total * (iso / total_partes)
        st.session_state["pedido_total"].append({
            "Código": codigo,
            "Modelo": modelo,
            "Talla": corrida,
            "Cantidad pares": cantidad,
            "Peso Total (g)": peso_total,
            "Poliol (g)": cantidad_poliol,
            "ISO (g)": cantidad_iso,
            "Hoja": ficha['Hoja']
        })
        st.success(f"Agregado: {modelo} - Talla {corrida} - {cantidad} pares.")
        st.session_state["corrida_seleccionada"] = None

if st.session_state["pedido_total"]:
    st.markdown("---")
    st.subheader("Resumen del Pedido")

    for idx, pedido in enumerate(st.session_state["pedido_total"]):
        col1, col2, col3, col4 = st.columns([4, 2, 2, 1])
        with col1:
            st.markdown(f"**Modelo:** {pedido['Modelo']} | **Talla:** {pedido['Talla']} | **Cantidad:** {pedido['Cantidad pares']} pares | **Hoja:** {pedido['Hoja']}")
        with col2:
            st.markdown(f"**Poliol:** {pedido['Poliol (g)']:.2f} g")
        with col3:
            st.markdown(f"**ISO:** {pedido['ISO (g)']:.2f} g")
        with col4:
            if st.button("Eliminar", key=f"eliminar_{idx}"):
                st.session_state["pedido_total"].pop(idx)
                st.rerun()

    resumen_df = pd.DataFrame(st.session_state["pedido_total"])
    if not resumen_df.empty:
        st.dataframe(resumen_df.style.format({
            "Peso Total (g)": "{:,.2f}",
            "Poliol (g)": "{:,.2f}",
            "ISO (g)": "{:,.2f}"
        }))

        total_poliol = resumen_df["Poliol (g)"].sum() / 1000
        total_iso = resumen_df["ISO (g)"].sum() / 1000
        mezcla_total_kg = total_poliol + total_iso
        merma_kg = mezcla_total_kg * 0.03
        mezcla_con_merma = mezcla_total_kg + merma_kg

        st.markdown(f"**Poliol necesario:** {total_poliol:.2f} kg")
        st.markdown(f"**ISO necesario:** {total_iso:.2f} kg")
        st.markdown(f"**Mezcla sin merma:** {mezcla_total_kg:.2f} kg")
        st.markdown(f"**Merma (3%):** {merma_kg:.2f} kg")
        st.markdown(f"**Mezcla total con merma:** {mezcla_con_merma:.2f} kg")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Exportar a Excel"):
                folder = Path("pedidos")
                folder.mkdir(exist_ok=True)
                fecha = datetime.now().strftime("%Y-%m-%d")
                codigos = "_".join(sorted(resumen_df["Código"].unique()))
                base_name = f"PEDIDO_SUOLMEX_{fecha}_{codigos}"
                nombre_archivo = folder / f"{base_name}.xlsx"
                contador = 1
                while nombre_archivo.exists():
                    nombre_archivo = folder / f"{base_name} ({contador}).xlsx"
                    contador += 1

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Pedido"
                for r in dataframe_to_rows(resumen_df, index=False, header=True):
                    ws.append(r)
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="003366")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max_length + 2

                ws2 = wb.create_sheet("Resumen")
                ws2["A1"] = "Resumen del Pedido"
                ws2["A1"].font = Font(size=14, bold=True)
                resumen = [
                    ("Total Poliol (kg):", total_poliol),
                    ("Total ISO (kg):", total_iso),
                    ("Mezcla sin merma (kg):", mezcla_total_kg),
                    ("Merma 3% (kg):", merma_kg),
                    ("Mezcla total con merma (kg):", mezcla_con_merma),
                    ("Versión del día:", f"{contador if contador > 1 else 1}")
                ]
                for i, (label, value) in enumerate(resumen, start=3):
                    ws2[f"A{i}"] = label
                    ws2[f"B{i}"] = value
                    ws2[f"A{i}"].font = Font(bold=True)
                    ws2[f"A{i}"].alignment = Alignment(horizontal="left")
                wb.save(nombre_archivo)
                st.success(f"Pedido exportado como: {nombre_archivo.name}")

        with col2:
            if st.button("Reiniciar Pedido"):
                st.session_state["pedido_total"] = []
                st.success("Pedido reiniciado.")

