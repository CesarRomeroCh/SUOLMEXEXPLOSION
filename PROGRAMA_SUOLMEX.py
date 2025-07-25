# PEDIDOS SUOLMEX CON LOGIN Y GESTIÓN DE USUARIOS
import os
import streamlit as st
import sqlite3
import pandas as pd
from datetime import datetime
import hashlib
import json
import uuid
from pathlib import Path
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración
DB_PATH = "usuarios.db"
FICHAS_PATH = "FICHAS2.xlsx"
LOGO_PATH = "logo_suolmex.jpg"
st.set_page_config(page_title="Calculadora de Pedido SUOLMEX", layout="wide")

# Estilo visual
st.markdown("""
    <style>
        body { background-color: #f5f7fa; }
        .stApp { font-family: 'Segoe UI', sans-serif; }
        h1, h2, h3 { color: #00264d; }
        .stButton>button {
            background-color: #00264d;
            color: white;
            font-weight: bold;
            padding: 8px 20px;
            border-radius: 6px;
            border: none;
        }
        .stButton>button:hover { background-color: #003366; }
        .stDataFrame thead tr th {
            background-color: #e6e9ef;
            color: #00264d;
        }
    </style>
""", unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image(LOGO_PATH, width=200)
    st.markdown("### Instrucciones")
    st.markdown("""
    1. Inicia sesión con tu código y contraseña.
    2. Los administradores pueden crear y editar usuarios.
    3. Los empleados pueden generar pedidos normalmente.
    """)

# Funciones de sesión
def obtener_session_id():
    if "session_id" not in st.session_state:
        st.session_state.session_id = uuid.uuid4().hex
    return st.session_state.session_id

def path_sesion_local():
    return f"session_{obtener_session_id()}.json"

def guardar_sesion():
    with open(path_sesion_local(), "w") as f:
        json.dump({
            "logueado": st.session_state.logueado,
            "usuario": st.session_state.usuario,
            "rol": st.session_state.rol
        }, f)

def cargar_sesion():
    try:
        with open(path_sesion_local(), "r") as f:
            data = json.load(f)
            st.session_state.logueado = data.get("logueado", False)
            st.session_state.usuario = data.get("usuario", None)
            st.session_state.rol = data.get("rol", None)
    except:
        st.session_state.logueado = False

# Conexión a DB
conn = sqlite3.connect(DB_PATH, check_same_thread=False)
c = conn.cursor()
c.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT UNIQUE NOT NULL,
        contrasena TEXT NOT NULL,
        rol TEXT NOT NULL
    )
""")
conn.commit()

def encriptar_contra(contra):
    return hashlib.sha256(contra.encode()).hexdigest()

# Crear admin si no existe
if not c.execute("SELECT * FROM usuarios WHERE codigo = 'admin'").fetchone():
    c.execute("INSERT INTO usuarios (codigo, contrasena, rol) VALUES (?, ?, ?)", 
              ('admin', encriptar_contra('admin123'), 'admin'))
    conn.commit()

# Cargar sesión
if "logueado" not in st.session_state:
    cargar_sesion()

# LOGIN
if not st.session_state.get("logueado", False):
    st.subheader("Iniciar sesión")
    with st.form("login_form"):
        codigo = st.text_input("Código de usuario")
        contrasena = st.text_input("Contraseña", type="password")
        login = st.form_submit_button("Entrar")
        if login:
            user = c.execute("SELECT contrasena, rol FROM usuarios WHERE codigo = ?", (codigo,)).fetchone()
            if user and encriptar_contra(contrasena) == user[0]:
                st.session_state.logueado = True
                st.session_state.usuario = codigo
                st.session_state.rol = user[1]
                guardar_sesion()
                st.rerun()
            else:
                st.error("Credenciales incorrectas.")
    st.stop()

# CERRAR SESIÓN
if st.button("Cerrar sesión"):
    ruta = path_sesion_local()
    if Path(ruta).exists(): Path(ruta).unlink()
    for k in list(st.session_state.keys()): del st.session_state[k]
    st.rerun()

# Bienvenida
st.success(f"Sesión iniciada como **{st.session_state.usuario}** ({st.session_state.rol})")

    # Historial de PDFs
with st.expander("Historial de pedidos generados"):
    folder = Path("historial_pedidos")
    if folder.exists():
         archivos_pdf = sorted(folder.glob("*.pdf"), reverse=True)
         if archivos_pdf:
             for archivo in archivos_pdf:
                with open(archivo, "rb") as f:
                    st.download_button(
                         label=f" {archivo.name}",
                         data=f,
                         file_name=archivo.name,
                         mime="application/pdf",
                          key=archivo.name
                       )
         else:
             st.info("No hay PDFs generados todavía.")
    else:
        st.info("No se ha generado ningún pedido aún.")

# GESTIÓN DE USUARIOS (solo admin)
if st.session_state.rol == "admin":
    st.markdown("---")
    st.subheader("Gestión de Usuarios")

    with st.expander("Crear nuevo usuario"):
        with st.form("crear_user"):
            nuevo = st.text_input("Nuevo código de usuario")
            contra = st.text_input("Contraseña", type="password")
            rol = st.selectbox("Rol", ["admin", "empleado"])
            if st.form_submit_button("Crear"):
                try:
                    c.execute("INSERT INTO usuarios (codigo, contrasena, rol) VALUES (?, ?, ?)",
                              (nuevo, encriptar_contra(contra), rol))
                    conn.commit()
                    st.success("Usuario creado.")
                except:
                    st.error("Ese código ya existe.")

    with st.expander("Editar o eliminar usuarios"):
        usuarios = pd.read_sql("SELECT codigo, rol FROM usuarios", conn)
        st.dataframe(usuarios)
        editar = st.selectbox("Selecciona usuario", usuarios["codigo"])
        nueva_contra = st.text_input("Nueva contraseña", type="password")
        if st.button("Actualizar contraseña"):
            c.execute("UPDATE usuarios SET contrasena = ? WHERE codigo = ?", 
                      (encriptar_contra(nueva_contra), editar))
            conn.commit()
            st.success("Contraseña actualizada.")
        if editar != "admin" and st.button("Eliminar usuario"):
            c.execute("DELETE FROM usuarios WHERE codigo = ?", (editar,))
            conn.commit()
            st.warning("Usuario eliminado.")

# FUNCIONALIDAD DEL SISTEMA DE PEDIDOS
st.markdown("---")
st.title("Generador de Pedido SUOLMEX")

@st.cache_data
def cargar_fichas(mtime: float):
    excel_file = pd.ExcelFile(FICHAS_PATH)
    hojas_deseadas = ['6001', '2066', '2060', '4098', 'PLANTILLAS']
    dataframes = []
    for hoja in hojas_deseadas:
        if hoja in excel_file.sheet_names:
            df_temp = excel_file.parse(hoja)
            df_temp['Hoja'] = hoja
            dataframes.append(df_temp)
    df = pd.concat(dataframes, ignore_index=True)
    df["Codigo del Producto"] = df["Codigo del Producto"].astype(str).str.strip()
    df["Linea"] = df["Linea"].astype(str).str.strip().str.upper()
    df["Corrida"] = df["Corrida"].astype(str).str.strip()
    df["Peso/Pie"] = pd.to_numeric(df["Peso/Pie"], errors="coerce")
    return df.dropna(subset=["Peso/Pie", "Relacion Poliol:ISO"])

# ———— Aquí pones el botón ————
if st.button("Recargar Excel de fichas"):
    st.cache_data.clear()
    st.experimental_rerun()

# ———— Y aquí cargas realmente las fichas ————
mtime = os.path.getmtime(FICHAS_PATH)
fichas = cargar_fichas(mtime)

# Estado
if "pedido_total" not in st.session_state:
    st.session_state["pedido_total"] = []
if "corrida_seleccionada" not in st.session_state:
    st.session_state["corrida_seleccionada"] = None

# Subir archivo
st.subheader("Subir archivo de pedido (opcional)")
archivo_pedido = st.file_uploader("Selecciona un archivo Excel con el pedido", type=["xlsx"])
if archivo_pedido:
    pedido_df = pd.read_excel(archivo_pedido)
    st.success("Pedido importado correctamente.")
    for _, row in pedido_df.iterrows():
        codigo = str(row["Codigo del Producto"]).strip()
        modelo = str(row["Modelo"]).strip().upper()
        corrida = str(row["Talla"]).strip()
        cantidad = int(row["Cantidad pares"])
        ficha = fichas[(fichas["Codigo del Producto"] == codigo) &
                       (fichas["Linea"] == modelo) &
                       (fichas["Corrida"] == corrida)]
        if not ficha.empty:
            ficha = ficha.iloc[0]
            peso_total = ficha['Peso/Pie'] * cantidad * 2
            poliol, iso = map(float, ficha['Relacion Poliol:ISO'].split(":"))
            total_partes = poliol + iso
            cantidad_poliol = peso_total * (poliol / total_partes)
            cantidad_iso = peso_total * (iso / total_partes)
            st.session_state["pedido_total"].append({
                "Código": codigo, "Modelo": modelo, "Talla": corrida, "Cantidad pares": cantidad,
                "Peso Total (g)": peso_total, "Poliol (g)": cantidad_poliol, "ISO (g)": cantidad_iso, "Hoja": ficha['Hoja']
            })

# Ingreso manual
st.markdown("---")
st.subheader("Ingreso manual")
codigo = st.selectbox("Código del Producto:", sorted(fichas["Codigo del Producto"].unique()))
modelos = fichas[fichas["Codigo del Producto"] == codigo]["Linea"].unique()
modelo = st.selectbox("Modelo:", sorted(modelos))
cantidad = st.number_input("Cantidad de pares:", min_value=1, step=1)
corridas = fichas[(fichas["Codigo del Producto"] == codigo) & (fichas["Linea"] == modelo)]["Corrida"].unique()
st.markdown("#### Selecciona una talla:")
cols = st.columns(min(5, len(corridas)))
for i, talla in enumerate(sorted(corridas)):
    if cols[i % 5].button(talla):
        st.session_state["corrida_seleccionada"] = talla

if st.session_state["corrida_seleccionada"]:
    corrida = st.session_state["corrida_seleccionada"]
    ficha = fichas[(fichas["Codigo del Producto"] == codigo) & (fichas["Linea"] == modelo) & (fichas["Corrida"] == corrida)]
    if not ficha.empty:
        ficha = ficha.iloc[0]
        peso_total = ficha['Peso/Pie'] * cantidad * 2
        poliol, iso = map(float, ficha['Relacion Poliol:ISO'].split(":"))
        total_partes = poliol + iso
        cantidad_poliol = peso_total * (poliol / total_partes)
        cantidad_iso = peso_total * (iso / total_partes)
        st.session_state["pedido_total"].append({
            "Código": codigo, "Modelo": modelo, "Talla": corrida, "Cantidad pares": cantidad,
            "Peso Total (g)": peso_total, "Poliol (g)": cantidad_poliol, "ISO (g)": cantidad_iso, "Hoja": ficha['Hoja']
        })
        st.success(f"Agregado: {modelo} - Talla {corrida} - {cantidad} pares.")
        st.session_state["corrida_seleccionada"] = None

# Resumen del pedido con opción de eliminar
if st.session_state["pedido_total"]:
    st.markdown("---")
    st.subheader("Resumen del Pedido")

    resumen_df = pd.DataFrame(st.session_state["pedido_total"])

    # Mostrar cada fila con botón de eliminación
    for idx, row in resumen_df.iterrows():
        cols = st.columns([5, 1])
        with cols[0]:
            st.markdown(
                f"**Código:** {row['Código']} | **Modelo:** {row['Modelo']} | "
                f"**Talla:** {row['Talla']} | **Cantidad:** {row['Cantidad pares']} pares | "
                f"**Poliol:** {row['Poliol (g)']:.2f} g | **ISO:** {row['ISO (g)']:.2f} g"
            )
        with cols[1]:
            if st.button("Eliminar", key=f"eliminar_{idx}"):
                st.session_state["pedido_total"].pop(idx)
                st.experimental_rerun()

    # Recalcular después de posibles eliminaciones
    resumen_df = pd.DataFrame(st.session_state["pedido_total"])
    total_poliol = resumen_df["Poliol (g)"].sum() / 1000
    total_iso = resumen_df["ISO (g)"].sum() / 1000
    poliol_merma = total_poliol * 0.03
    iso_merma = total_iso * 0.03
    total_poliol_con_merma = total_poliol + poliol_merma
    total_iso_con_merma = total_iso + iso_merma
    mezcla_total_kg = total_poliol + total_iso
    mezcla_con_merma = total_poliol_con_merma + total_iso_con_merma

    # Mostrar resultados
    st.markdown(f"*Poliol necesario:* {total_poliol:.2f} kg")
    st.markdown(f"*ISO necesario:* {total_iso:.2f} kg")
    st.markdown(f"*Poliol con merma (3%):* {total_poliol_con_merma:.2f} kg")
    st.markdown(f"*ISO con merma (3%):* {total_iso_con_merma:.2f} kg")
    st.markdown(f"*Mezcla sin merma:* {mezcla_total_kg:.2f} kg")
    st.markdown(f"*Mezcla total con merma:* {mezcla_con_merma:.2f} kg")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Generar PDF"):
            Path("historial_pedidos").mkdir(exist_ok=True)
            fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            nombre_usuario = st.session_state.usuario
            nombre_archivo = f"historial_pedidos/pedido_{nombre_usuario}_{fecha_hora}.pdf"

            class PDF(FPDF):
                def header(self):
                    self.image(LOGO_PATH, 10, 8, 33)
                    self.set_font("Arial", 'B', 12)
                    self.cell(0, 10, "Resumen de Pedido SUOLMEX", 0, 1, 'C')
                    self.ln(10)

            pdf = PDF()
            pdf.add_page()
            pdf.set_font("Arial", size=10)

            pdf.cell(0, 10, f"Usuario: {nombre_usuario}", ln=True)
            pdf.cell(0, 10, f"Fecha y hora: {fecha_hora.replace('_', ' ')}", ln=True)
            pdf.ln(5)

            headers = ["Código", "Modelo", "Talla", "Pares", "Poliol (g)", "ISO (g)"]
            for h in headers:
                pdf.cell(32, 10, h, 1, 0, 'C')
            pdf.ln()

            for _, row in resumen_df.iterrows():
                pdf.cell(32, 10, str(row['Código']), 1)
                pdf.cell(32, 10, str(row['Modelo']), 1)
                pdf.cell(32, 10, str(row['Talla']), 1)
                pdf.cell(32, 10, str(row['Cantidad pares']), 1)
                pdf.cell(32, 10, f"{row['Poliol (g)']:.1f}", 1)
                pdf.cell(32, 10, f"{row['ISO (g)']:.1f}", 1)
                pdf.ln()

            pdf.ln(5)
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 10, "Totales:", ln=True)
            pdf.set_font("Arial", size=10)
            pdf.cell(0, 10, f"Poliol necesario: {total_poliol:.2f} kg", ln=True)
            pdf.cell(0, 10, f"ISO necesario: {total_iso:.2f} kg", ln=True)
            pdf.cell(0, 10, f"Poliol con merma (3%): {total_poliol_con_merma:.2f} kg", ln=True)
            pdf.cell(0, 10, f"ISO con merma (3%): {total_iso_con_merma:.2f} kg", ln=True)
            pdf.cell(0, 10, f"Mezcla sin merma: {mezcla_total_kg:.2f} kg", ln=True)
            pdf.cell(0, 10, f"Mezcla total con merma: {mezcla_con_merma:.2f} kg", ln=True)

            pdf.output(nombre_archivo)
            st.success(f"PDF generado: {Path(nombre_archivo).name}")

            with open(nombre_archivo, "rb") as f:
                st.download_button("📥 Descargar PDF", data=f, file_name=Path(nombre_archivo).name)

    with col2:
        if st.button("Reiniciar Pedido"):
            st.session_state["pedido_total"] = []
            st.success("Pedido reiniciado.")

